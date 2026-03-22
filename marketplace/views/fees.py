"""
Thống kê phí sàn (Fees) - Báo cáo doanh thu và phí sàn từ giao dịch.
- statistics: tổng revenue, phí, tx + theo N ngày (?days=7|14|30)
- top_transactions: 5 giao dịch có phí cao nhất
- save: lưu snapshot báo cáo (report_type=FEES)
- export: xuất báo cáo phí sàn ra file (mở được bằng Excel)
"""
from datetime import timedelta
import csv
import io
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from django.db.models import Sum, Count
from django.http import HttpResponse
from django.utils import timezone
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAdminUser
from rest_framework.response import Response

from ..models import Transaction, AdminReportSnapshot
from ..serializers import TransactionListSerializer


@api_view(['POST'])
@permission_classes([IsAdminUser])
def save_fees_report(request):
    """Lưu báo cáo phí sàn: stats và timeseries vào AdminReportSnapshot."""
    stats_data = request.data.get('stats') or {}
    timeseries_data = request.data.get('timeseries') or {}
    
    # 2. Lưu vào Database (snapshot)
    AdminReportSnapshot.objects.create(
        report_type='FEES',
        snapshot_data={'stats': stats_data, 'timeseries': timeseries_data},
        created_by=request.user,
    )

    return Response({'status': 'saved', 'message': 'Đã lưu snapshot báo cáo phí sàn.'})


@api_view(['GET'])
@permission_classes([IsAdminUser])
def fee_statistics(request):
    """GET ?days=N: Tổng revenue, phí, tx + revenue/phí N ngày gần nhất, phí TB/giao dịch."""
    days = int(request.query_params.get('days', 7))
    if days < 1:
        days = 7
    if days > 90:
        days = 90

    now = timezone.now()
    since = now - timedelta(days=days)

    summary = Transaction.objects.aggregate(
        total_revenue=Sum('amount'),
        total_platform_fee=Sum('platform_fee'),
        total_transactions=Count('id'),
    )
    last_n = Transaction.objects.filter(created_at__gte=since).aggregate(
        rev=Sum('amount'),
        fee=Sum('platform_fee'),
    )
    rev_n = float(last_n['rev'] or 0)
    fee_n = float(last_n['fee'] or 0)
    cnt = summary['total_transactions'] or 0
    fee_total = summary['total_platform_fee'] or 0
    avg_fee = float(fee_total / cnt) if cnt else 0

    data = {
        'total_revenue': summary['total_revenue'],
        'total_platform_fee': summary['total_platform_fee'],
        'total_transactions': summary['total_transactions'],
        'revenue_last_n_days': rev_n,
        'platform_fee_last_n_days': fee_n,
        'avg_fee_per_transaction': avg_fee,
        'days': days,
    }
    return Response(data)


@api_view(['GET'])
@permission_classes([IsAdminUser])
def fee_top_transactions(request):
    """Lấy 5 giao dịch có phí sàn cao nhất."""
    qs = Transaction.objects.select_related('listing', 'buyer').order_by('-platform_fee')[:5]
    serializer = TransactionListSerializer(qs, many=True)
    return Response(serializer.data)


@api_view(['GET'])
@permission_classes([IsAdminUser])
def export_fees_report_csv(request):
    """
    Xuất báo cáo phí sàn ra file CSV để tải về (mở bằng Excel).

    File gồm:
    - Tổng quan phí sàn (fee_statistics)
    - Timeseries (doanh thu & phí sàn theo ngày)
    - Top 5 giao dịch có phí cao nhất
    """
    days = int(request.query_params.get('days', 7))
    if days < 1:
        days = 7
    if days > 90:
        days = 90

    now = timezone.now()
    since = now - timedelta(days=days)
    today = timezone.localdate()
    start_date = today - timedelta(days=days - 1)

    summary = Transaction.objects.aggregate(
        total_revenue=Sum('amount'),
        total_platform_fee=Sum('platform_fee'),
        total_transactions=Count('id'),
    )
    last_n = Transaction.objects.filter(created_at__gte=since).aggregate(
        rev=Sum('amount'),
        fee=Sum('platform_fee'),
    )
    rev_n = float(last_n['rev'] or 0)
    fee_n = float(last_n['fee'] or 0)
    cnt = summary['total_transactions'] or 0
    fee_total = summary['total_platform_fee'] or 0
    avg_fee = float(fee_total / cnt) if cnt else 0

    txs_qs = (
        Transaction.objects.filter(created_at__date__gte=start_date)
        .values('created_at')
        .annotate(revenue=Sum('amount'), fee=Sum('platform_fee'))
    )
    labels = [(start_date + timedelta(days=i)).isoformat() for i in range(days)]
    tx_map = {
        row['created_at'].date().isoformat(): {
            'revenue': float(row['revenue'] or 0),
            'fee': float(row['fee'] or 0),
        }
        for row in txs_qs
    }

    top_qs = Transaction.objects.select_related('listing', 'buyer').order_by('-platform_fee')[:5]
    top_serializer = TransactionListSerializer(top_qs, many=True)

    filename = f"fees-report-{today.isoformat()}"
    
    # Khởi tạo workbook Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Fees Report"

    # Định dạng font cơ bản
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True)

    # Phần 1: Summary
    ws.append(['BÁO CÁO PHÍ SÀN'])
    ws['A1'].font = title_font
    ws.append(['Số ngày', days])
    ws.append([])
    ws.append(['Chỉ số', 'Giá trị'])
    # In đậm hàng tiêu đề
    for cell in ws[4]:
        cell.font = header_font

    ws.append(['Tổng doanh thu', float(summary['total_revenue'] or 0)])
    ws.append(['Tổng phí sàn', float(summary['total_platform_fee'] or 0)])
    ws.append(['Tổng giao dịch', summary['total_transactions'] or 0])
    ws.append([f'Doanh thu {days} ngày gần nhất', rev_n])
    ws.append([f'Phí sàn {days} ngày gần nhất', fee_n])
    ws.append(['Phí trung bình mỗi giao dịch', avg_fee])

    # Đổ data theo ngày (Timeseries)
    ws.append([])
    ws.append(['DỮ LIỆU THEO NGÀY'])
    ws.cell(row=ws.max_row, column=1).font = title_font
    ws.append(['Ngày', 'Doanh thu', 'Phí sàn'])
    for cell in ws[ws.max_row]:
        cell.font = header_font

    for label in labels:
        tx_data = tx_map.get(label, {})
        ws.append([
            label,
            tx_data.get('revenue', 0),
            tx_data.get('fee', 0),
        ])

    # Top 5 giao dịch có phí cao nhất
    ws.append([])
    ws.append(['TOP 5 GIAO DỊCH CÓ PHÍ CAO NHẤT'])
    ws.cell(row=ws.max_row, column=1).font = title_font
    ws.append(['ID', 'Bài đăng', 'Người mua', 'Số tiền', 'Phí sàn', 'Ngày'])
    for cell in ws[ws.max_row]:
        cell.font = header_font

    for row in top_serializer.data:
        ws.append([
            row['id'],
            row['listing_title'],
            row['buyer_username'],
            float(row['amount']),
            float(row['platform_fee']),
            row['created_at'],
        ])

    # Auto fix độ rộng cột cho đẹp
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    wb.save(response)
    
    return response
