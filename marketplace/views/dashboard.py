"""
Bảng điều khiển (Dashboard) - Thống kê tổng quan hệ thống.
- summary: 6 thẻ số liệu (user, listing, transaction, revenue, ...)
- timeseries: dữ liệu theo ngày cho biểu đồ tăng trưởng
- save: lưu snapshot báo cáo (AdminReportSnapshot)
- export: xuất báo cáo ra file (mở được bằng Excel)
"""
from datetime import timedelta
import csv
import io
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from django.contrib.auth.models import User
from django.db.models import Sum, Count
from django.db.models.functions import TruncDate
from django.http import HttpResponse
from django.utils import timezone
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAdminUser
from rest_framework.response import Response

from ..models import Listing, Transaction, AdminReportSnapshot, UserReport

@api_view(['GET'])
@permission_classes([IsAdminUser])
def admin_notifications(request):
    """
    API cung cấp số lượng thông báo cho Admin (dành cho đồ án).
    - Đếm số bài đăng đang chờ duyệt (PENDING)
    - Đếm số báo cáo người dùng đang mở (OPEN)
    """
    # Đếm các bài đăng sản phẩm chưa được duyệt
    pending_listings_count = Listing.objects.filter(status='PENDING').count()
    
    # Đếm các báo cáo tố cáo người dùng chưa được xử lý
    open_reports_count = UserReport.objects.filter(status='OPEN').count()
    
    # Tạo danh sách các thông báo cụ thể để hiển thị trên menu chuông
    notifications = []
    
    if pending_listings_count > 0:
        notifications.append({
            'title': 'Bài đăng mới',
            'message': f'Có {pending_listings_count} bài đăng đang chờ bạn phê duyệt.',
            'time': 'Vừa xong',
            'type': 'LISTING',
            'link': '/pages/listings'
        })
        
    if open_reports_count > 0:
        notifications.append({
            'title': 'Báo cáo vi phạm',
            'message': f'Có {open_reports_count} báo cáo cần xem xét ngay.',
            'time': 'Vừa xong',
            'type': 'REPORT',
            'link': '/pages/reports'
        })

    return Response({
        'unread_count': pending_listings_count + open_reports_count,
        'items': notifications
    })

@api_view(['POST'])
@permission_classes([IsAdminUser])
def save_dashboard_report(request):
    """Lưu báo cáo dashboard: summary và timeseries vào AdminReportSnapshot."""
    summary_data = request.data.get('summary') or {}
    timeseries_data = request.data.get('timeseries') or {}
    
    # 2. Lưu vào Database (giữ lại snapshot)
    AdminReportSnapshot.objects.create(
        report_type='DASHBOARD',
        snapshot_data={'summary': summary_data, 'timeseries': timeseries_data},
        created_by=request.user,
    )

    return Response({'status': 'saved', 'message': 'Đã lưu hệ thống (snapshot) thành công.'})


@api_view(['GET'])
@permission_classes([IsAdminUser])
def export_dashboard_report_csv(request):
    """
    Xuất báo cáo dashboard ra file CSV để tải về (mở bằng Excel).

    File gồm 2 phần:
    - Tổng quan (summary)
    - Dữ liệu theo ngày (timeseries)
    """
    days = int(request.query_params.get('days', 7))
    if days < 1:
        days = 7
    if days > 90:
        days = 90

    now = timezone.now()
    since = now - timedelta(days=days)
    today = timezone.localdate()
    start_date = today - timedelta(days=days - 1)

    total_users = User.objects.count()
    total_listings = Listing.objects.count()
    total_transactions = Transaction.objects.count()
    total_revenue = Transaction.objects.aggregate(total=Sum('amount'))['total'] or 0
    listings_last_n_days = Listing.objects.filter(created_at__gte=since).count()
    transactions_last_n_days = Transaction.objects.filter(created_at__gte=since).count()

    listings_qs = (
        Listing.objects.filter(created_at__date__gte=start_date)
        .annotate(d=TruncDate('created_at'))
        .values('d')
        .annotate(c=Count('id'))
    )
    txs_qs = (
        Transaction.objects.filter(created_at__date__gte=start_date)
        .annotate(d=TruncDate('created_at'))
        .values('d')
        .annotate(c=Count('id'), revenue=Sum('amount'), fee=Sum('platform_fee'))
    )
    labels = [(start_date + timedelta(days=i)).isoformat() for i in range(days)]
    listing_map = {row['d'].isoformat(): row['c'] for row in listings_qs}
    tx_map = {
        row['d'].isoformat(): {
            'count': row['c'],
            'revenue': float(row['revenue'] or 0),
            'fee': float(row['fee'] or 0),
        }
        for row in txs_qs
    }

    filename = f"dashboard-report-{today.isoformat()}"
    
    # Khởi tạo file Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard Report"

    # Style cho font và căn lề
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True)
    center_align = Alignment(horizontal='center')

    # Đổ dữ liệu tổng quan (Summary)
    ws.append(['BÁO CÁO DASHBOARD'])
    ws['A1'].font = title_font
    ws.append(['Số ngày', days])
    ws.append([])
    ws.append(['Chỉ số', 'Giá trị'])
    # In đậm header
    for cell in ws[4]: 
        cell.font = header_font

    ws.append(['Tổng người dùng', total_users])
    ws.append(['Tổng bài đăng', total_listings])
    ws.append(['Tổng giao dịch', total_transactions])
    ws.append(['Tổng doanh thu', float(total_revenue)])
    ws.append([f'Bài đăng {days} ngày gần nhất', listings_last_n_days])
    ws.append([f'Giao dịch {days} ngày gần nhất', transactions_last_n_days])

    # Đổ dữ liệu chi tiết theo ngày (Timeseries)
    ws.append([])
    ws.append(['DỮ LIỆU THEO NGÀY'])
    start_ts_row = ws.max_row
    ws.cell(row=start_ts_row, column=1).font = title_font
    
    ws.append(['Ngày', 'Số bài đăng mới', 'Số giao dịch', 'Doanh thu', 'Phí sàn'])
    for cell in ws[ws.max_row]:
        cell.font = header_font

    for label in labels:
        listing_count = listing_map.get(label, 0)
        tx_data = tx_map.get(label, {})
        ws.append([
            label,
            listing_count,
            tx_data.get('count', 0),
            tx_data.get('revenue', 0),
            tx_data.get('fee', 0),
        ])

    # Tự động căn chỉnh độ rộng cột cho dễ nhìn
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    wb.save(response)
    
    return response


@api_view(['GET'])
@permission_classes([IsAdminUser])
def dashboard_summary(request):
    """Thống kê tổng quan: 6 thẻ (người dùng, bài đăng, giao dịch, doanh thu, bài đăng N ngày, giao dịch N ngày)."""
    days = int(request.query_params.get('days', 7))
    if days < 1:
        days = 7
    if days > 90:
        days = 90

    now = timezone.now()
    since = now - timedelta(days=days)

    total_users = User.objects.count()
    total_listings = Listing.objects.count()
    total_transactions = Transaction.objects.count()
    total_revenue = Transaction.objects.aggregate(total=Sum('amount'))['total'] or 0

    listings_last_n_days = Listing.objects.filter(created_at__gte=since).count()
    transactions_last_n_days = Transaction.objects.filter(created_at__gte=since).count()

    return Response({
        'total_users': total_users,
        'total_listings': total_listings,
        'total_transactions': total_transactions,
        'total_revenue': total_revenue,
        'listings_last_n_days': listings_last_n_days,
        'transactions_last_n_days': transactions_last_n_days,
        'days': days,
    })


@api_view(['GET'])
@permission_classes([IsAdminUser])
def dashboard_data(request):
    """Gộp summary và timeseries trong một request để load nhanh hơn."""
    days = int(request.query_params.get('days', 7))
    if days < 1:
        days = 7
    if days > 90:
        days = 90

    now = timezone.now()
    since = now - timedelta(days=days)
    today = timezone.localdate()
    start_date = today - timedelta(days=days - 1)

    total_users = User.objects.count()
    total_listings = Listing.objects.count()
    total_transactions = Transaction.objects.count()
    total_revenue = Transaction.objects.aggregate(total=Sum('amount'))['total'] or 0
    listings_last_n_days = Listing.objects.filter(created_at__gte=since).count()
    transactions_last_n_days = Transaction.objects.filter(created_at__gte=since).count()

    labels = [(start_date + timedelta(days=i)).isoformat() for i in range(days)]
    listings_qs = (
        Listing.objects.filter(created_at__date__gte=start_date)
        .annotate(d=TruncDate('created_at'))
        .values('d')
        .annotate(c=Count('id'))
    )
    txs_qs = (
        Transaction.objects.filter(created_at__date__gte=start_date)
        .annotate(d=TruncDate('created_at'))
        .values('d')
        .annotate(c=Count('id'), revenue=Sum('amount'), fee=Sum('platform_fee'))
    )
    listing_map = {row['d'].isoformat(): row['c'] for row in listings_qs}
    tx_map = {
        row['d'].isoformat(): {
            'count': row['c'],
            'revenue': float(row['revenue'] or 0),
            'fee': float(row['fee'] or 0),
        }
        for row in txs_qs
    }

    return Response({
        'summary': {
            'total_users': total_users,
            'total_listings': total_listings,
            'total_transactions': total_transactions,
            'total_revenue': total_revenue,
            'listings_last_n_days': listings_last_n_days,
            'transactions_last_n_days': transactions_last_n_days,
            'days': days,
        },
        'timeseries': {
            'labels': labels,
            'listings_created': [listing_map.get(d, 0) for d in labels],
            'transactions_count': [tx_map.get(d, {}).get('count', 0) for d in labels],
            'revenue': [tx_map.get(d, {}).get('revenue', 0) for d in labels],
            'platform_fee': [tx_map.get(d, {}).get('fee', 0) for d in labels],
        },
    })


@api_view(['GET'])
@permission_classes([IsAdminUser])
def dashboard_timeseries(request):
    """Dữ liệu theo ngày cho biểu đồ: số bài đăng mới, giao dịch, doanh thu, phí sàn mỗi ngày."""
    days = int(request.query_params.get('days', 7))
    if days < 1:
        days = 7
    if days > 90:
        days = 90

    today = timezone.localdate()
    start_date = today - timedelta(days=days - 1)
    labels = [(start_date + timedelta(days=i)).isoformat() for i in range(days)]

    listings = (
        Listing.objects.filter(created_at__date__gte=start_date)
        .annotate(d=TruncDate('created_at'))
        .values('d')
        .annotate(c=Count('id'))
    )
    txs = (
        Transaction.objects.filter(created_at__date__gte=start_date)
        .annotate(d=TruncDate('created_at'))
        .values('d')
        .annotate(c=Count('id'), revenue=Sum('amount'), fee=Sum('platform_fee'))
    )

    listing_map = {row['d'].isoformat(): row['c'] for row in listings}
    tx_map = {
        row['d'].isoformat(): {
            'count': row['c'],
            'revenue': float(row['revenue'] or 0),
            'fee': float(row['fee'] or 0),
        }
        for row in txs
    }

    return Response({
        'labels': labels,
        'listings_created': [listing_map.get(d, 0) for d in labels],
        'transactions_count': [tx_map.get(d, {}).get('count', 0) for d in labels],
        'revenue': [tx_map.get(d, {}).get('revenue', 0) for d in labels],
        'platform_fee': [tx_map.get(d, {}).get('fee', 0) for d in labels],
    })
